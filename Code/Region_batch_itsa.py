import os
from pathlib import Path
import pandas as pd
import numpy as np
import statsmodels.api as sm
import matplotlib.pyplot as plt
from matplotlib.patches import Rectangle
from statsmodels.tsa.stattools import adfuller
from statsmodels.stats.stattools import durbin_watson
from statsmodels.stats.diagnostic import het_breuschpagan
from scipy.stats import zscore, mannwhitneyu, norm
import sys
import importlib.util

lib_dir = Path(__file__).resolve().parents[2] / 'lib'
itsa_utils_path = lib_dir / 'itsa_utils.py'
spec = importlib.util.spec_from_file_location('itsa_utils', itsa_utils_path)
if spec is None or spec.loader is None:
    raise ImportError(f'Cannot load utilities module from {itsa_utils_path}')
itsa_utils = importlib.util.module_from_spec(spec)
spec.loader.exec_module(itsa_utils)

compute_mean_ci = itsa_utils.compute_mean_ci
compute_mean_hac_ci = itsa_utils.compute_mean_hac_ci
fit_sarimax_with_retry = itsa_utils.fit_sarimax_with_retry
block_bootstrap_median_diff = itsa_utils.block_bootstrap_median_diff
parametric_bootstrap_cf_ci = itsa_utils.parametric_bootstrap_cf_ci
run_diagnostics_and_choose_se = itsa_utils.run_diagnostics_and_choose_se


BASE_DATA_PATH = r'D:\检测小队\林嘉意\中断时间序列\Global\data\PFOS_NAH.csv'
NP_DATA_PATH = r'D:\检测小队\林嘉意\中断时间序列\Global\data\PFOS_NAH_NP.csv'


def ensure_dir(p):
    Path(p).mkdir(parents=True, exist_ok=True)


def compute_mean_ci_with_used(model, frame, choice: str = 'HAC', maxlags: int = 4, alpha: float = 0.05):
    """Compute mean prediction CI like itsa_utils.compute_mean_ci but also return which covariance was actually used.

    Returns: (pred_mean, lower, upper, used_method)
    """
    pred_mean = model.predict(frame)
    choice_up = (choice or 'OLS').upper()
    used = None
    try:
        if choice_up == 'HAC':
            robust = itsa_utils._robust_results(model, cov_type='HAC', maxlags=maxlags)
            cov = robust.cov_params()
            used = 'HAC'
        elif choice_up == 'HC3':
            robust = itsa_utils._robust_results(model, cov_type='HC3')
            cov = robust.cov_params()
            used = 'HC3'
        else:
            cov = model.cov_params()
            used = 'OLS'
        cov_arr = cov.values if hasattr(cov, 'values') else np.asarray(cov)
    except Exception:
        cov_arr = model.cov_params().values if hasattr(model.cov_params(), 'values') else np.asarray(model.cov_params())
        used = 'OLS_fallback'

    exog = itsa_utils._prediction_exog(model, frame)
    se_mean = np.sqrt(np.einsum('ij,jk,ik->i', exog, cov_arr, exog))
    z = -1 * norm.ppf(alpha / 2)
    lower = pred_mean - z * se_mean
    upper = pred_mean + z * se_mean
    return pred_mean, lower, upper, used


def compute_medians(df):
    df['semester'] = df['month'].apply(lambda x: 'H1' if x <= 6 else 'H2')
    median_raw = df.groupby(['year', 'semester', 'month', 'source', 'lon', 'lat', 'lon_grid', 'lat_grid'])['value [ng/L]'].median().reset_index()
    median_cell = median_raw.groupby(['year', 'semester', 'month', 'source', 'lon_grid', 'lat_grid'])['value [ng/L]'].median().reset_index()
    median_ref = median_cell.groupby(['year', 'semester', 'month', 'lon_grid', 'lat_grid'])['value [ng/L]'].median().reset_index()
    median_month = median_ref.groupby(['year', 'semester', 'month'])['value [ng/L]'].median().reset_index()
    median_year_semester = median_month.groupby(['year', 'semester'])['value [ng/L]'].median().reset_index()
    return median_ref, median_year_semester


def prepare_timeseries(median_year_semester):
    df = median_year_semester.copy()
    df['year_semester'] = (df['year'].astype(str) + '-' + df['semester'].replace({'H1': '01-01', 'H2': '07-01'}))
    df['year_semester'] = pd.to_datetime(df['year_semester'])
    df = df.sort_values('year_semester').drop_duplicates('year_semester')
    df['time'] = np.arange(len(df))
    return df


def detect_and_remove_outliers(df):
    obs = ~df['value [ng/L]'].isna()
    X = sm.add_constant(df.loc[obs, 'time'])
    y = df.loc[obs, 'value [ng/L]']
    model = sm.OLS(y, X).fit()
    df.loc[obs, 'resid'] = model.resid
    df.loc[obs, 'z'] = zscore(df.loc[obs, 'resid'])
    df['is_outlier'] = False
    df.loc[obs, 'is_outlier'] = df.loc[obs, 'z'].abs() > 3
    return df


def save_removed_outliers(df: pd.DataFrame, out_path: Path, region_name: str) -> pd.DataFrame:
    removed = df.loc[df['is_outlier']].copy()
    columns = [c for c in ['year', 'month', 'year_month', 'time', 'value [ng/L]', 'resid', 'z', 'is_outlier'] if c in removed.columns]
    removed = removed[columns]
    removed.to_csv(out_path, index=False, encoding='utf-8-sig')
    print(f"[{region_name}] 已导出异常值清单到：{out_path}")
    print(f"[{region_name}] 删除的异常值数量：{len(removed)}")
    return removed


def choose_final_cov_type(bp_info: dict, dw_info: dict) -> str:
    """Choose OLS / HC3 / HAC based on heteroscedasticity and autocorrelation checks."""
    if dw_info.get('has_autocorrelation', False):
        return 'HAC'
    if bp_info.get('has_heteroscedasticity', False):
        return 'HC3'
    return 'OLS'


def get_final_model_results(model, cov_type: str, maxlags: int = 4):
    if cov_type == 'HC3':
        return model.get_robustcov_results(cov_type='HC3')
    if cov_type == 'HAC':
        return model.get_robustcov_results(cov_type='HAC', maxlags=maxlags)
    return model


def interpolate_full_series(df):
    full_idx = pd.date_range(start=df['year_semester'].min(), end=df['year_semester'].max(), freq='6MS')
    data = pd.DataFrame({'year_semester': full_idx})
    data = data.merge(df[['year_semester', 'value [ng/L]', 'observed']], on='year_semester', how='left')
    data['value [ng/L]'] = data['value [ng/L]'].interpolate(method='linear')
    data['value [ng/L]'] = data['value [ng/L]'].bfill().ffill()
    data['year'] = data['year_semester'].dt.year
    data['month'] = data['year_semester'].dt.month
    data['year_month'] = data['year_semester']
    data = data.reset_index(drop=True)
    return data


def fit_itsa_and_save(region_dir, data, intervention_date, run_label: str = ''):
    data['time'] = np.arange(len(data))
    data['period'] = (data['year_month'] >= pd.to_datetime(intervention_date)).astype(int)
    T0 = data[data['year_month'] == pd.to_datetime(intervention_date)]['time'].values[0]
    data['time_c'] = data['time'] - T0
    # time_after for compatibility with SARIMAX/3_ITSA_SARIMAX implementation
    data['time_after'] = np.where(data['period'] == 1, data['time'] - T0, 0)

    model1 = sm.OLS.from_formula('Q("value [ng/L]") ~ time + period', data=data).fit()
    model2 = sm.OLS.from_formula('Q("value [ng/L]") ~ time + time:period', data=data).fit()
    model3 = sm.OLS.from_formula('Q("value [ng/L]") ~ time + period + I(time-{0}):period'.format(T0), data=data).fit()

    # Print OLS summaries to terminal so user can inspect coefficients
    print(f"\n[{region_dir}] OLS Model summaries:")
    print('\n--- Model 1: Level Change Only ---')
    print(model1.summary())
    print('\n--- Model 2: Slope Change Only ---')
    print(model2.summary())
    print('\n--- Model 3: Level + Slope Change ---')
    print(model3.summary())

    # Diagnostics and automatic SE choice
    diag1 = run_diagnostics_and_choose_se(model1, name='Model1: Level Change Only')
    diag2 = run_diagnostics_and_choose_se(model2, name='Model2: Slope Change Only')
    diag3 = run_diagnostics_and_choose_se(model3, name='Model3: Level + Slope Change')

    # Print robust SEs for transparency
    for name, model in [('Model1', model1), ('Model2', model2), ('Model3', model3)]:
        try:
            hc3 = model.get_robustcov_results(cov_type='HC3')
            nw = model.get_robustcov_results(cov_type='HAC', maxlags=4)
            print(f"\n{name} HC3 SE:\n", hc3.bse)
            print(f"{name} HAC SE:\n", nw.bse)
        except Exception:
            print(f"\n{name} robust SE unavailable")

    def get_preds(m):
        pred = m.predict(data)
        data_cf = data.copy()
        data_cf['period'] = 0
        pred_cf = m.predict(data_cf)
        return pred, pred_cf

    pred1, pred_cf1 = get_preds(model1)
    pred2, pred_cf2 = get_preds(model2)
    pred3, pred_cf3 = get_preds(model3)

    # diagnostics
    pre_data = data[data['year_month'] < pd.to_datetime(intervention_date)]['value [ng/L]']
    adf_pre = adfuller(pre_data) if len(pre_data.dropna()) > 3 else (np.nan, np.nan)

    # save results
    results = pd.DataFrame({
        'Model': ['Level Change Only', 'Slope Change Only', 'Level + Slope Change'],
        'AIC': [model1.aic, model2.aic, model3.aic],
        'Intervention Effect (Level)': [model1.params.get('period', np.nan), model2.params.get('period', np.nan), model3.params.get('period', np.nan)],
        'Intervention Effect (Slope)': [model1.params.get('time_c:period', np.nan), model2.params.get('time_c:period', np.nan), model3.params.get('time_c:period', np.nan)],
        'Chosen_SE': [diag1['chosen_se'], diag2['chosen_se'], diag3['chosen_se']],
        'BP_pvalue': [diag1['bp_pvalue'], diag2['bp_pvalue'], diag3['bp_pvalue']],
        'DW': [diag1['dw'], diag2['dw'], diag3['dw']],
        'BG_pvalue': [diag1['bg_pvalue'], diag2['bg_pvalue'], diag3['bg_pvalue']],
        'LjungBox_pvalue': [diag1['lb_pvalue'], diag2['lb_pvalue'], diag3['lb_pvalue']],
    })
    # results will be saved later after determining which CI was actually used for Model3

    # Use diagnostic-chosen covariance for mean CI (HAC/HC3/OLS) and record actual method used
    # Determine chosen SE for Model3
    model3_choice = diag3['chosen_se']
    pred_mean, lower, upper, used_ci = compute_mean_ci_with_used(model3, data, choice=model3_choice, maxlags=4)
    data_cf = data.copy()
    data_cf['period'] = 0
    data_cf['time'] = data['time']
    data_cf['time_after'] = 0
    pred_cf_mean, lower_cf, upper_cf, used_cf_ci = compute_mean_ci_with_used(model3, data_cf, choice=model3_choice, maxlags=4)

    print(f"\n[{region_dir}] Final SE choice by model:")
    print(results[['Model', 'Chosen_SE', 'BP_pvalue', 'DW', 'BG_pvalue', 'LjungBox_pvalue']].to_string(index=False))
    print(f"\n[{region_dir}] Model3 selected SE: {model3_choice}")
    print(f"[{region_dir}] Actual CI method used for Model3 fitted series: {used_ci}")
    print(f"[{region_dir}] Actual CI method used for Model3 counterfactual: {used_cf_ci}")
    print_paper_ready_results(model3, model3_choice, maxlags=4)

    # parametric bootstrap CF CI
    try:
        ci_cf_low_boot, ci_cf_high_boot = parametric_bootstrap_cf_ci(model3, data_cf, nboot=500)
    except Exception:
        ci_cf_low_boot, ci_cf_high_boot = lower_cf, upper_cf

    # include which CI method was actually used (e.g., HAC/HC3/OLS)
    fitted_data = pd.DataFrame({
        'year_month': data['year_month'],
        'Model 3 Fitted': np.asarray(pred_mean),
        'Model 3 CI Lower': np.asarray(lower),
        'Model 3 CI Upper': np.asarray(upper),
        'Model 3 Counterfactual': np.asarray(pred_cf_mean),
        'Model 3 CF CI Lower': np.asarray(lower_cf),
        'Model 3 CF CI Upper': np.asarray(upper_cf),
        'Model 3 CF CI Lower (Boot)': np.asarray(ci_cf_low_boot),
        'Model 3 CF CI Upper (Boot)': np.asarray(ci_cf_high_boot),
        'CI_Method': [used_ci] * len(data),
        'CF_CI_Method': [used_cf_ci] * len(data),
    })
    fitted_path = os.path.join(region_dir, f'model3_fitted_data{run_label}.xlsx')
    fitted_data.to_excel(fitted_path, index=False)

    # record which CI method was actually used for Model3 and save results summary
    try:
        results['Model3_Used_CI'] = [np.nan, np.nan, used_ci]
    except Exception:
        results['Model3_Used_CI'] = [None, None, used_ci]
    results.to_csv(os.path.join(region_dir, f'PFOS_ITSA_Results{run_label}.csv'), index=False)
    save_paper_ready_results(model3, model3_choice, Path(region_dir), 'PFOS_ITSA', suffix=run_label.strip('_'), maxlags=4)

    # --- SARIMAX sensitivity (fit SARIMAX(1,0,1) as in main script) ---
    try:
        print(f"\n[{region_dir}] Fitting SARIMAX(1,0,1) sensitivity...")
        endog = data['value [ng/L]']
        exog = data[['time', 'period', 'time_after']]
        sarimax_res = fit_sarimax_with_retry(endog, exog, (1, 0, 1), label='Sensitivity')
        print(f"SARIMAX summary for {region_dir}:")
        print(sarimax_res.summary())

        pred = sarimax_res.get_prediction(start=0, end=len(data) - 1, exog=exog)
        sarimax_frame = data.copy()
        sarimax_frame['sarimax_fitted'] = pred.predicted_mean
        ci = pred.conf_int()
        ci_arr = ci.to_numpy() if hasattr(ci, 'to_numpy') else np.asarray(ci)
        sarimax_frame['sarimax_ci_lower'] = ci_arr[:, 0]
        sarimax_frame['sarimax_ci_upper'] = ci_arr[:, 1]

        sarimax_out_csv = os.path.join(region_dir, f'sarimax_sensitivity_fitted{run_label}.csv')
        sarimax_frame.to_csv(sarimax_out_csv, index=False)
        print(f"Saved SARIMAX fitted series to: {sarimax_out_csv}")

        # save sarimax plot
        try:
            plt.figure(figsize=(10, 5))
            observed = sarimax_frame[sarimax_frame['value [ng/L]'].notna()]
            plt.scatter(observed['year_month'], observed['value [ng/L]'], s=20, color='black', label='Observed')
            plt.plot(sarimax_frame['year_month'], sarimax_frame['sarimax_fitted'], color='#1f77b4', lw=2, label='SARIMAX fitted')
            plt.fill_between(sarimax_frame['year_month'], sarimax_frame['sarimax_ci_lower'], sarimax_frame['sarimax_ci_upper'], color='#1f77b4', alpha=0.18)
            plt.axvline(pd.to_datetime(intervention_date), color='crimson', ls='--', lw=1.2, label='Intervention')
            plt.title('SARIMAX Sensitivity')
            plt.xlabel('Year')
            plt.ylabel('PFOS Concentration [ng/L]')
            plt.legend()
            plt.tight_layout()
            plt.savefig(os.path.join(region_dir, f'SARIMAX_Sensitivity{run_label}.png'))
            plt.close()
        except Exception:
            pass
    except Exception as e:
        print(f"SARIMAX sensitivity failed for {region_dir}: {e}")

    # plot and save
    try:
        plt.figure(figsize=(10, 5))
        ax = plt.gca()
        intervention_idx = data.index[data['year_month'] == pd.to_datetime(intervention_date)][0]
        ax.add_patch(Rectangle((intervention_idx, 0), len(data) - intervention_idx, data['value [ng/L]'].max() * 1.2, color='grey', alpha=0.2))
        observed_data = data[data['observed'] == 1]
        plt.scatter(observed_data.index, observed_data['value [ng/L]'], s=20, label='Observed')
        # use diagnostic-chosen CI (computed earlier) instead of model default conf_int
        plt.plot(data.index, pred3, 'b-', lw=2, label='Fitted')
        plt.fill_between(data.index, lower, upper, color='blue', alpha=0.2)
        plt.plot(data.index, pred_cf3, 'r--', lw=2, label='Counterfactual')
        plt.fill_between(data.index, lower_cf, upper_cf, color='red', alpha=0.15)
        plt.title('Model 3 Fitted vs Counterfactual')
        plt.legend()
        plt.tight_layout()
        plt.savefig(os.path.join(region_dir, f'model3_fitted{run_label}.png'))
        plt.close()
    except Exception:
        pass

    # Return fitted models so caller can aggregate OLS results if desired
    return {'model1': model1, 'model2': model2, 'model3': model3}


def save_ols_results_to_excel(models: dict, workbook_path: str, sheet_prefix: str):
    """Save OLS coefficient tables for models to an Excel workbook.
    Creates or appends sheets named '<sheet_prefix>_Model1', etc.
    """
    from pathlib import Path
    workbook_exists = Path(workbook_path).exists()
    try:
        # pandas >= 1.3 supports if_sheet_exists
        mode = 'a' if workbook_exists else 'w'
        with pd.ExcelWriter(workbook_path, engine='openpyxl', mode=mode, if_sheet_exists='replace') as writer:
            for name, m in models.items():
                try:
                    ci = m.conf_int()
                except Exception:
                    ci = pd.DataFrame()
                df = pd.DataFrame({
                    'coef': m.params,
                    'bse': m.bse,
                    't': m.tvalues,
                    'pvalue': m.pvalues,
                })
                if not ci.empty:
                    ci.columns = ['ci_low', 'ci_high']
                    df = df.join(ci)

                # compute HAC robust results (HAC with maxlags=4)
                try:
                    robust = m.get_robustcov_results(cov_type='HAC', maxlags=4)
                    hac_ci = robust.conf_int()
                    if hasattr(hac_ci, 'columns'):
                        hac_ci.columns = ['HAC_ci_low', 'HAC_ci_high']
                        df = df.join(hac_ci)
                    # add robust bse and pvalues
                    df['HAC_bse'] = robust.bse
                    # robust may expose pvalues
                    try:
                        df['HAC_pvalue'] = robust.pvalues
                    except Exception:
                        df['HAC_pvalue'] = np.nan
                except Exception:
                    df['HAC_bse'] = np.nan
                    df['HAC_pvalue'] = np.nan
                sheet_name = f"{sheet_prefix}_{name}"
                # Excel sheet name max length is 31
                sheet_name = sheet_name[:31]
                df.to_excel(writer, sheet_name=sheet_name)
    except TypeError:
        # Fallback for older pandas that don't support if_sheet_exists
        with pd.ExcelWriter(workbook_path, engine='openpyxl', mode='a' if workbook_exists else 'w') as writer:
            for name, m in models.items():
                try:
                    ci = m.conf_int()
                except Exception:
                    ci = pd.DataFrame()
                df = pd.DataFrame({
                    'coef': m.params,
                    'bse': m.bse,
                    't': m.tvalues,
                    'pvalue': m.pvalues,
                })
                if not ci.empty:
                    ci.columns = ['ci_low', 'ci_high']
                    df = df.join(ci)
                try:
                    robust = m.get_robustcov_results(cov_type='HAC', maxlags=4)
                    hac_ci = robust.conf_int()
                    if hasattr(hac_ci, 'columns'):
                        hac_ci.columns = ['HAC_ci_low', 'HAC_ci_high']
                        df = df.join(hac_ci)
                    df['HAC_bse'] = robust.bse
                    try:
                        df['HAC_pvalue'] = robust.pvalues
                    except Exception:
                        df['HAC_pvalue'] = np.nan
                except Exception:
                    df['HAC_bse'] = np.nan
                    df['HAC_pvalue'] = np.nan
                sheet_name = f"{sheet_prefix}_{name}"
                sheet_name = sheet_name[:31]
                df.to_excel(writer, sheet_name=sheet_name)


def save_model_summaries_to_excel(models: dict, workbook_path: str, sheet_prefix: str):
    """Save full model.summary().as_text() into Excel workbook.
    Each model gets a sheet named '<sheet_prefix>_<model>_summary' with the summary text written one line per row in column A.
    """
    from pathlib import Path
    try:
        from openpyxl import load_workbook, Workbook
    except Exception:
        raise

    p = Path(workbook_path)
    if p.exists():
        wb = load_workbook(workbook_path)
    else:
        wb = Workbook()
        # remove default sheet if empty
        if 'Sheet' in wb.sheetnames and len(wb.sheetnames) == 1:
            std = wb['Sheet']
            wb.remove(std)

    for name, m in models.items():
        sheet_name = f"{sheet_prefix}_{name}_summary"[:31]
        if sheet_name in wb.sheetnames:
            # replace existing sheet
            ws = wb[sheet_name]
            wb.remove(ws)
        ws = wb.create_sheet(sheet_name)
        try:
            summary_text = m.summary().as_text()
        except Exception:
            summary_text = str(m.summary())
        for i, line in enumerate(summary_text.splitlines(), start=1):
            ws.cell(row=i, column=1, value=line)

    wb.save(workbook_path)


def _display_param_name(name: str) -> str:
    if name == 'Intercept':
        return 'Intercept'
    if name == 'time':
        return 'time'
    if name == 'period':
        return 'period'
    if name == 'time_after':
        return 'time_after'
    if 'I(time -' in name and ':period' in name:
        return 'time_after'
    return name


def paper_ready_parameter_table(model, cov_type: str, maxlags: int = 4) -> pd.DataFrame:
    final_res = get_final_model_results(model, cov_type=cov_type, maxlags=maxlags)
    param_index = list(model.params.index)
    params = pd.Series(np.asarray(final_res.params), index=param_index)
    bse = pd.Series(np.asarray(final_res.bse), index=param_index)
    pvalues = pd.Series(np.asarray(final_res.pvalues), index=param_index)
    conf = pd.DataFrame(np.asarray(final_res.conf_int()), index=param_index)

    rows = []
    for raw_name in param_index:
        display_name = _display_param_name(raw_name)
        if display_name not in {'Intercept', 'time', 'period', 'time_after'}:
            continue
        rows.append({
            'Parameter': display_name,
            'Estimate': params.get(raw_name, np.nan),
            'SE': bse.get(raw_name, np.nan),
            'p_value': pvalues.get(raw_name, np.nan),
            'CI_low': conf.loc[raw_name, 0] if raw_name in conf.index else np.nan,
            'CI_high': conf.loc[raw_name, 1] if raw_name in conf.index else np.nan,
        })
    return pd.DataFrame(rows)


def print_paper_ready_results(model, cov_type: str, maxlags: int = 4) -> None:
    paper_df = paper_ready_parameter_table(model, cov_type=cov_type, maxlags=maxlags)
    print('\n=== 论文可直接使用的最终参数结果 ===')
    print(f'最终采用的标准误类型：{cov_type}')
    print(paper_df.round(4).to_string(index=False))


def save_paper_ready_results(model, cov_type: str, output_dir: Path, file_stem: str, suffix: str = '', maxlags: int = 4) -> pd.DataFrame:
    paper_df = paper_ready_parameter_table(model, cov_type=cov_type, maxlags=maxlags)
    output_dir.mkdir(parents=True, exist_ok=True)
    label = f'_{suffix}' if suffix else ''
    csv_path = output_dir / f'{file_stem}{label}_final_parameters_{cov_type}.csv'
    xlsx_path = output_dir / f'{file_stem}{label}_final_parameters_{cov_type}.xlsx'
    paper_df.to_csv(csv_path, index=False, encoding='utf-8-sig')
    paper_df.to_excel(xlsx_path, index=False)
    print(f'\n最终参数表已保存到：{csv_path}')
    print(f'最终参数表已保存到：{xlsx_path}')
    return paper_df


def run(base_data_path=BASE_DATA_PATH, output_root=None):
    print(f"[ITSA] Input CSV: {base_data_path}", flush=True)
    df_all = pd.read_csv(base_data_path, dtype={'value [ng/L]': float})
    df_all['source'] = df_all['source'].fillna('Unknown')
    df_all['limit [ng/L]'] = df_all['limit [ng/L]'].fillna(0.003536)
    df_all.loc[df_all['value [ng/L]'] == 0, 'value [ng/L]'] = np.nan
    df_all = df_all[~((df_all['limit [ng/L]'] >= 1) & (df_all['value [ng/L]'].isna()))].copy()
    missing_value_mask = df_all['value [ng/L]'].isna()
    df_all.loc[missing_value_mask, 'value [ng/L]'] = df_all.loc[missing_value_mask, 'limit [ng/L]'] / np.sqrt(2)
    #  df1.loc[missing_value_mask, 'value [ng/L]'] = 0.003536 / np.sqrt(2)
    df_all['country'] = df_all['country'].astype(str).str.upper()

    regions = [
        {'name': '美国', 'filter': lambda d: d['country'] == 'UNITED STATES', 'cutoff': '2007-10-01', 'intervention': '2007-07-01'},
        {'name': '欧盟', 'filter': None, 'cutoff': '2010-08-01', 'intervention': '2010-07-01'},
        {'name': '中国', 'filter': lambda d: d['country'] == 'CHINA', 'cutoff': '2014-04-01', 'intervention': '2014-01-01'},
    ]

    EU_countries = ["AUSTRIA", "BELGIUM", "BULGARIA", "CROATIA", "CYPRUS", "CZECHIA", "DENMARK", "ESTONIA", "FINLAND",
                    "FRANCE", "GERMANY", "HUNGARY", "IRELAND", "ITALY", "LATVIA", "LITHUANIA", "LUXEMBOURG", "MALTA",
                    "NETHERLANDS", "POLAND", "PORTUGAL", "ROMANIA", "SLOVAKIA", "SLOVENIA", "GREECE", "SPAIN", "SWEDEN",
                    "UNITED KINGDOM"]

    if output_root is None:
        base_out = Path(__file__).parent.resolve() / 'data_processing'
    else:
        base_out = Path(output_root)
    print(f"[ITSA] Output root: {base_out}", flush=True)
    for reg in regions:
        name = reg['name']
        region_dir = base_out / name
        print(f"\n[ITSA] Processing region: {name}", flush=True)
        print(f"[ITSA] Region output dir: {region_dir}", flush=True)
        ensure_dir(region_dir)
        if name == '欧盟':
            df = df_all[df_all['country'].isin(EU_countries)].copy()
        else:
            df = df_all[df_all['country'] == ("UNITED STATES" if name == '美国' else 'CHINA')].copy()

        df['source'] = df['source'].fillna('Unknown')
        median_ref, median_year_semester = compute_medians(df)
        median_ref.to_csv(region_dir / f"{name}_median_cell_PFOS.csv", index=False)
        median_year_semester.to_csv(region_dir / f"{name}_semester_median_PFOS.csv", index=False)

        # prepare timeseries + outliers
        ts = prepare_timeseries(median_year_semester)
        ts = detect_and_remove_outliers(ts)
        save_removed_outliers(ts, region_dir / f'{name}_removed_outliers.csv', name)
        ts.loc[~ts['is_outlier'], 'observed'] = 1
        cleaned = ts[~ts['is_outlier']].copy()
        cleaned['observed'] = cleaned['observed'].fillna(0)

        data = interpolate_full_series(cleaned)
        data.to_csv(region_dir / 'semester_clean_once_interp.csv', index=False)

        # If region is United States, run multiple intervention semester points
        if name == '美国':
            # list of semester labels to run (keep original 2007H2 and add new intervention points)
            us_semesters = ['2007H2', '2009H2', '2015H1', '2016H1', '2020H2', '2021H1', '2021H2', '2022H1']
            # master workbook to collect OLS regression results
            master_workbook = os.path.join(region_dir, 'US_multiple_interventions_OLS_results.xlsx')
            for sem in us_semesters:
                # convert semester label to intervention date string
                try:
                    year = int(sem[:4])
                    half = sem[4:]
                except Exception:
                    continue
                if half == 'H1':
                    intervention_date = f"{year}-01-01"
                else:
                    intervention_date = f"{year}-07-01"

                run_label = f"_{sem}"
                print(f"[ITSA] Running US intervention {sem} -> {intervention_date}", flush=True)
                models = fit_itsa_and_save(str(region_dir), data.copy(), intervention_date, run_label=run_label)
                try:
                    save_ols_results_to_excel(models, master_workbook, sem)
                except Exception as e:
                    print(f"Failed to save OLS results for {sem}: {e}", flush=True)
        else:
            fit_itsa_and_save(str(region_dir), data, reg['intervention'])


if __name__ == '__main__':
    input_csv = os.environ.get('ITSA_INPUT_CSV', BASE_DATA_PATH)
    output_root = os.environ.get('ITSA_OUTPUT_ROOT')
    explicit_input = os.environ.get('ITSA_INPUT_CSV')

    # If ITSA_INPUT_CSV is explicitly provided, run only that dataset.
    if explicit_input:
        run(input_csv, output_root)
    else:
        # Default behavior: run baseline PFOS_NAH first, then NP sensitivity dataset.
        print('\n[ITSA] Running baseline dataset (PFOS_NAH.csv)...', flush=True)
        run(BASE_DATA_PATH, output_root)

        run_np = os.environ.get('ITSA_RUN_NP', '1').strip().lower() not in {'0', 'false', 'no'}
        np_input = os.environ.get('ITSA_NP_INPUT_CSV', NP_DATA_PATH)
        np_output = os.environ.get(
            'ITSA_NP_OUTPUT_ROOT',
            str(Path(__file__).parent.resolve() / 'data_processing' / 'No Point')
        )
        if run_np:
            if Path(np_input).exists():
                print('\n[ITSA] Running additional dataset (PFOS_NAH_NP.csv)...', flush=True)
                run(np_input, np_output)
            else:
                print(f"[ITSA] Skip PFOS_NAH_NP run: file not found -> {np_input}", flush=True)
